import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import re
import os

# Katalog wejściowy
input_dir = "Input"
publications_list = [
    os.path.join(input_dir, file)
    for file in os.listdir(input_dir)
    if file.endswith(".xlsx")
]


def safe_get(dataframe, row, col, default=""):
    try:
        return dataframe.iloc[row, col]
    except IndexError:
        return default


def extract_first_author(authors):
    # ekstrakcja pierwszego autora do zapisania nazwy pliku
    # Podzielić według pierwszego wystąpienia 'and' lub ','
    first_author = re.split(r" and |, ", authors)[0]
    # Usunąć wszystko po pierwszej spacji, łącznie ze spacją
    first_author = first_author.split(" ")[0]
    return first_author


def make_title(title) -> str:
    title = re.sub(r'[<>:"/\\|?*]', "", title)  # Usunięcie niedozwolonych znaków
    title = re.sub(
        r"[\[\]{}().\-]", "", title
    )  # Usunięcie nawiasów, kropek i myślników
    title = re.sub(r"\s+", "-", title)  # Zamiana wielu spacji na pojedynczą "-"
    # ogarniecie tytułu do zapisu pliku
    lngh = 50  # dlugosc titleu
    idx = [i for i, sym in enumerate(title) if sym == "-"]
    for e, _ in enumerate(idx):
        if len(title) < lngh:
            break
        else:
            if len(title[: idx[e]]) < lngh:
                continue
            title = (
                title[: idx[e]] if len(title[: idx[e]]) >= lngh else title[: idx[e - 1]]
            )
            break
    return title


def create_xlsx_files(dir_file):
    df = pd.read_excel(dir_file)
    endpoint = os.path.splitext(os.path.basename(dir_file))[0]
    for row in df.index.values:
        authors = safe_get(df, row, 10)
        first_author = extract_first_author(authors)
        doi = safe_get(df, row, 17)
        year = safe_get(df, row, 2)
        title = safe_get(df, row, 1)
        filename = f"{first_author}_{year}_" + make_title(title) + f"_{endpoint}"

        # WCZYTAJ PLIK
        workbook = load_workbook("ECHA_NAM_DATA_TEMPLATE.xlsx")
        # WCZYTAJ SHEET
        worksheet = workbook["SCORING"]
        print(workbook.sheetnames)
        input()

        # DODANIE LISTY DROPDOWN DO ARKUSZA PO UKRYTYM DICTIONARY
        dictionary_sheet = workbook["endpoints_dictionary"]
        dictionary_range = "A2:A26"
        dv = DataValidation(
            type="list",
            formula1=f"='{dictionary_sheet.title}'!{dictionary_range}",
            showDropDown=False,
        )
        dv.prompt = "Please select a endpoint from the list"
        dv.promptTitle = "Endpoints List"
        worksheet.add_data_validation(dv)
        dv.add(worksheet["E3"])

        # WPISANIE WARTOSCI DO ARKUSZA
        insert = [("A3", authors), ("B3", title), ("C3", doi)]
        for cell, value in insert:
            worksheet[cell] = value

        # ZAPISANIE PLIKU Z ODPOWIEDNIA NAZWA
        directory = "./Output"
        if not os.path.exists(directory):
            # If it doesn't exist, create it
            os.makedirs(directory)
        else:
            if not os.path.exists(os.path.join(directory, endpoint)):
                os.makedirs(os.path.join(directory, endpoint))
        workbook.save(f"{directory}/{endpoint}/{filename}.xlsx")


[create_xlsx_files(f) for f in publications_list]
