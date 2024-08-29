import pandas as pd
from openpyxl import load_workbook
import re


def extract_first_author(authors):
    # ekstrakcja pierwszego autora do zapisania nazwy pliku
    # Podzielić według pierwszego wystąpienia 'and' lub ','
    first_author = re.split(r" and |, ", authors)[0]
    # Usunąć wszystko po pierwszej spacji, łącznie ze spacją
    first_author = first_author.split(" ")[0]
    return first_author


def make_title(title) -> str:
    title = re.sub(r'[<>:"/\\|?*]', "", title)  # Usunięcie niedozwolonych znaków
    title = re.sub(
        r"[\[\]{}().\-]", "", title
    )  # Usunięcie nawiasów, kropek i myślników
    title = re.sub(r"\s+", "-", title)  # Zamiana wielu spacji na pojedynczą "-"
    # ogarniecie tytułu do zapisu pliku
    lngh = 50  # dlugosc titleu
    idx = [i for i, sym in enumerate(title) if sym == "-"]
    for e, _ in enumerate(idx):
        if len(title) < lngh:
            break
        else:
            if len(title[: idx[e]]) < lngh:
                continue
            title = (
                title[: idx[e]] if len(title[: idx[e]]) >= lngh else title[: idx[e - 1]]
            )
            break
    return title


def take_methods(df):
    result = []
    columns = df.columns
    for _, row in df.iterrows():
        all_columns = row.values.copy()
        col20_value = all_columns[20]

        if isinstance(col20_value, str):
            elements = col20_value.split("\n")
            for element in elements:
                new_row = list(all_columns)
                new_row[20] = element
                result.append(new_row)
        else:
            result.append(all_columns)
    return pd.DataFrame(result, columns=columns)


def fill_missing_values(df):
    last_valid_row = None
    for index, row in df.iterrows():
        if pd.notna(row["No"]) and pd.notna(row["Authors"]):
            last_valid_row = row
        elif pd.isna(row["Authors"]):
            if last_valid_row is not None:
                for col in df.columns:
                    if pd.isna(row[col]):
                        df.at[index, col] = last_valid_row[col]
    return df


def fill_missing_values2(df):
    last_values = {}

    for index, row in df.iterrows():
        author = row["Authors"]

        if author not in last_values:
            last_values[author] = {}

        for col in df.columns:
            if col != "Authors" and pd.isna(row[col]):
                if col in last_values[author]:
                    df.at[index, col] = last_values[author][col]
            else:
                last_values[author][col] = row[col]

    return df


def remove_zeros(df, start=0, end=8):
    last_values = {}

    for index, row in df.iterrows():
        author = row["Authors"]

        if author not in last_values:
            last_values[author] = {}

        for col in df.columns[start:end]:
            if col != "Authors" and row[col] == 0:
                if col in last_values[author]:
                    df.at[index, col] = last_values[author][col]
            else:
                last_values[author][col] = row[col]

    df.iloc[:, start:end] = df.iloc[:, start:end].replace(0, pd.NA)
    return df


def save_author_files(df: pd.DataFrame):
    unique_authors = df["Authors"].unique()
    for author in unique_authors:
        df_author = df[df["Authors"] == author]
        first_author = extract_first_author(author)
        title = df_author["Title"].iloc[0]
        filename = f"{first_author}_" + make_title(title) + ".xlsx"
        if df_author.iloc[:, 10:].notna().any().any():
            path2save = f"./TO_REMOVE/WITH_SCORING/{filename}"
            df_author.to_excel(path2save, index=False, sheet_name="NAMS")
        else:
            wb = load_workbook("../ECHA_TEMPLATE_TASK/ECHA_NAM_DATA_TEMPLATE.xlsx")
            sheets = wb.sheetnames  # ['Sheet1', 'Sheet2']

            for s in sheets:
                if s == "NAMs":
                    sheet_name = wb[s]
                    wb.remove(sheet_name)
            path2save = f"./TO_REMOVE/WITHOUT_SCORING/{filename}"
            wb.save(path2save)
            df_author.iloc[:, 8] = "='S&K'!F4"
            df_author.iloc[:, 9] = "='S&K'!F8"
            with pd.ExcelWriter(path2save, engine="openpyxl", mode="a") as writer:
                df_author.to_excel(writer, index=False, sheet_name="NAMS")


def cleaner_chain(inventory: pd.DataFrame, filename: str):
    result_df = take_methods(inventory)
    result_df = result_df.dropna(subset=list(result_df.columns[10:]), how="all")
    result_df = fill_missing_values(result_df)
    result_df = fill_missing_values2(result_df)
    result_df = remove_zeros(result_df)
    return result_df.to_excel(filename, index=False)


def pipleine2excels(inventory: pd.DataFrame):
    result_df = take_methods(inventory)
    result_df = fill_missing_values(result_df)
    result_df = fill_missing_values2(result_df)
    result_df = remove_zeros(result_df)
    result_df = result_df.dropna(subset=list(result_df.columns[0:7]), how="all")
    return save_author_files(result_df)
